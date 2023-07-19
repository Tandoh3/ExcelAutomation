import openpyxl

from openpyxl import Workbook, load_workbook 
from openpyxl.utils import get_column_letter
'''
#loading an existing workbook 
wb= load_workbook("Grades.xlsx")

#Accessing the Sheet 
ws = wb.active
'''
#print(ws)

#accessing an ndividual cell value 
#print(ws['A2'].value)

#change the value 
#ws["A2"].value = "Manuel"

#save workbook so that the change can take effect 
#wb.save("Grades.xlsx")

#Creating, Listing and Changing Sheets
#print(wb.sheetnames)
'''

#accessing sheets 
#print(wb["Sheet1"])

#create a new sheet 
wb.create_sheet("Test")
print(wb.sheetnames)
wb.save("Grades.xlsx")
'''

'''
#creating a new workbook 
wb = Workbook() 
ws = wb.active
ws.title = "Data"

#adding and appending 
ws.append(["Tim", "Is", "Great", "!"])
ws.append(["Tim", "Is", "Great", "!"])
ws.append(["Tim", "Is", "Great", "!"])
ws.append(["Tim", "Is", "Great", "!"])
ws.append(["end"])

#save workbook 
wb.save('tim.xlsx')

'''


'''
#Acessing Multiple Cells 
wb = load_workbook("tim.xlsx")
ws = wb.active

for row in range(1, 11):
    for col in range(1, 5):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)
'''

'''
#merging cells 
wb = load_workbook("tim.xlsx")
ws = wb.active

ws.merge_cells("A1:D1")
ws.unmerge_cells("A1:D1")

wb.save("tim.xlsx")
'''

#insert empty rows 
wb = load_workbook("tim.xlsx")
ws = wb.active



    
    

